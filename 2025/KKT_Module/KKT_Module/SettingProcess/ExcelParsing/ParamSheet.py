
import time
import typing
import openpyxl as pxl
from KKT_Module.KKT_Module.SettingProcess.ExcelParsing.Core import Register, SymbolChar, BitRow, Sheet, RevisionSheet, FilePathSheet, HWSettingSheet

class ParamSheet():
    def __init__(self, file_name:str):
        self.file_name = file_name
        self.sheet_names = []
        self.workbook:typing.Dict[str, Sheet] = {}
    def __getitem__(self, item):
        return self.__getattribute__(item)

    def get(self, item, ret=None):
        return self[item]
    def parseKsocExcel(self, wb:pxl.Workbook):
        sheetnames = wb.sheetnames
        for name in sheetnames:
            name = name.strip()
            if name.find('$') == 0:
                self.sheet_names.append(name)
                if name.find("$RevisionHistory") == 0:
                    sheet = RevisionSheet()
                elif name.find("$AI_WeightData") == 0:
                    sheet = FilePathSheet("$AI_WeightData")
                else:
                    sheet = HWSettingSheet(name=name)
                self.workbook[sheet.name] = sheet
                sheet.parseSheet(wb[name])

    def asDict(self)->dict:
        d = {}
        work_book = {}
        for k, v in self.__dict__.items():
            if k == 'workbook':
                for name, sheet in self.workbook.items():
                    work_book.update({name: sheet.asDict()})
                    v = work_book
            d.setdefault(k, v)
        return d


if __name__ == '__main__':
    from pprint import pprint
    from pathlib import Path
    from KKT_Module.KKT_Module.ksoc_global import kgl
    import json
    file = str(Path(kgl.KKTModule).joinpath(r'SettingProcess\Test_Param\168A_Format\kkt60SOCA1_2D_SIC_RX13_32chirp_Dongle.xlsx'))
    ps = ParamSheet(file_name=file)
    s = time.time_ns()
    wb = pxl.load_workbook(file)
    ps.parseKsocExcel(wb)
    wb.close()
    print(ps.workbook)
    print(f'{(time.time_ns()-s)/1000000000} s')

    with open('HW_setting.json', 'w') as outfile:
        json.dump(ps.asDict(), outfile, ensure_ascii=False)
        outfile.write('\n')
    print('')



